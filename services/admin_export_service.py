from __future__ import annotations

from datetime import date, datetime, time
from decimal import Decimal
from io import BytesIO
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session, joinedload, selectinload

from models.booking import Booking
from models.booking_assignment import BookingAssignment
from models.cleaner_profile import CleanerProfile
from models.payment import Payment
from models.rating import Rating
from models.role import Role
from models.user import User
from models.user_role import UserRole


EXPORT_DATASETS = {"all", "cleaners", "bookings", "ratings", "users", "payments"}


def build_admin_export_workbook(db: Session, dataset: str) -> BytesIO:
    normalized_dataset = dataset.strip().lower()
    if normalized_dataset not in EXPORT_DATASETS:
        raise ValueError("Unsupported export dataset")

    workbook = Workbook()
    workbook.remove(workbook.active)

    datasets = (
        ["users", "cleaners", "bookings", "ratings", "payments"]
        if normalized_dataset == "all"
        else [normalized_dataset]
    )

    builders = {
        "users": _build_users_sheet,
        "cleaners": _build_cleaners_sheet,
        "bookings": _build_bookings_sheet,
        "ratings": _build_ratings_sheet,
        "payments": _build_payments_sheet,
    }

    for name in datasets:
        headers, rows = builders[name](db)
        _append_sheet(workbook, name.title(), headers, rows)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def export_filename(dataset: str) -> str:
    timestamp = datetime.utcnow().strftime("%Y%m%d-%H%M%S")
    return f"washioo-{dataset.strip().lower()}-export-{timestamp}.xlsx"


def _build_users_sheet(db: Session):
    users = (
        db.query(User)
        .options(joinedload(User.user_roles).joinedload(UserRole.role))
        .order_by(User.created_at.desc())
        .all()
    )
    headers = [
        "User ID",
        "Full Name",
        "Phone",
        "Email",
        "Roles",
        "Verified",
        "Active",
        "Terms Accepted",
        "Average Rating",
        "Total Ratings",
        "Last Login",
        "Created At",
        "Updated At",
    ]
    rows = [
        [
            user.id,
            user.full_name,
            user.phone,
            user.email,
            _role_names(user),
            user.is_verified,
            user.is_active,
            user.terms_accepted,
            user.average_rating,
            user.total_ratings,
            user.last_login,
            user.created_at,
            user.updated_at,
        ]
        for user in users
    ]
    return headers, rows


def _build_cleaners_sheet(db: Session):
    cleaners = (
        db.query(CleanerProfile)
        .options(
            joinedload(CleanerProfile.user),
            joinedload(CleanerProfile.document_reviewer),
        )
        .order_by(CleanerProfile.created_at.desc())
        .all()
    )
    headers = [
        "Cleaner Profile ID",
        "User ID",
        "Full Name",
        "Phone",
        "Email",
        "Vehicle Type",
        "Approval Status",
        "Availability Status",
        "Verification Status",
        "Document Review Status",
        "Document Rejection Reason",
        "Resubmission Required",
        "Documents Submitted At",
        "Documents Verified At",
        "Documents Reviewed By",
        "Profile Photo URL",
        "Aadhaar Image URL",
        "Driving License Image URL",
        "Service Radius KM",
        "Auto Assign Enabled",
        "Current Latitude",
        "Current Longitude",
        "Last Location At",
        "Last Available At",
        "Average Rating",
        "Total Ratings",
        "Jobs Completed",
        "Created At",
        "Updated At",
    ]
    rows = [
        [
            cleaner.id,
            cleaner.user_id,
            cleaner.user.full_name if cleaner.user else None,
            cleaner.user.phone if cleaner.user else None,
            cleaner.user.email if cleaner.user else None,
            cleaner.vehicle_type,
            cleaner.approval_status,
            cleaner.availability_status,
            cleaner.verification_status,
            cleaner.document_review_status,
            cleaner.document_rejection_reason,
            cleaner.document_resubmission_required,
            cleaner.documents_submitted_at,
            cleaner.documents_verified_at,
            cleaner.document_reviewer.full_name if cleaner.document_reviewer else None,
            cleaner.profile_photo_url,
            cleaner.aadhaar_image_url,
            cleaner.driving_license_image_url,
            cleaner.service_radius_km,
            cleaner.auto_assign_enabled,
            cleaner.current_latitude,
            cleaner.current_longitude,
            cleaner.last_location_at,
            cleaner.last_available_at,
            cleaner.average_rating,
            cleaner.total_ratings,
            cleaner.total_jobs_completed,
            cleaner.created_at,
            cleaner.updated_at,
        ]
        for cleaner in cleaners
    ]
    return headers, rows


def _build_bookings_sheet(db: Session):
    bookings = (
        db.query(Booking)
        .options(
            selectinload(Booking.customer),
            selectinload(Booking.service_category),
            selectinload(Booking.address),
            selectinload(Booking.vehicle),
            selectinload(Booking.assignment)
            .selectinload(BookingAssignment.cleaner)
            .selectinload(CleanerProfile.user),
            selectinload(Booking.payment),
        )
        .order_by(Booking.created_at.desc())
        .all()
    )
    headers = [
        "Booking ID",
        "Booking Reference",
        "Customer ID",
        "Customer Name",
        "Customer Phone",
        "Customer Email",
        "Service",
        "Scheduled Date",
        "Scheduled Time",
        "Booking Status",
        "Estimated Price",
        "Final Price",
        "Vehicle Make",
        "Vehicle Model",
        "License Plate",
        "Address Line 1",
        "Address Line 2",
        "Landmark",
        "City",
        "State",
        "Pincode",
        "Latitude",
        "Longitude",
        "Cleaner ID",
        "Cleaner Name",
        "Assignment Status",
        "Assigned At",
        "Accepted At",
        "Started At",
        "Completed At",
        "Payment Status",
        "Legacy Payment Status",
        "Payment Type",
        "Collected Amount",
        "Cleaner Share",
        "Admin Share",
        "Cleaner Handover Status",
        "Special Instructions",
        "Created At",
        "Updated At",
    ]
    rows = []
    for booking in bookings:
        address = booking.address
        assignment = booking.assignment
        cleaner = assignment.cleaner if assignment else None
        payment = booking.payment
        rows.append(
            [
                booking.id,
                booking.booking_reference,
                booking.customer_id,
                booking.customer.full_name if booking.customer else None,
                booking.customer.phone if booking.customer else None,
                booking.customer.email if booking.customer else None,
                booking.service_category.service_name if booking.service_category else None,
                booking.scheduled_date,
                booking.scheduled_time,
                booking.booking_status,
                booking.estimated_price,
                booking.final_price,
                booking.vehicle_make,
                booking.vehicle_model,
                booking.license_plate,
                address.address_line1 if address else None,
                address.address_line2 if address else None,
                address.landmark if address else None,
                address.city if address else None,
                address.state if address else None,
                address.pincode if address else None,
                address.latitude if address else None,
                address.longitude if address else None,
                cleaner.id if cleaner else None,
                cleaner.user.full_name if cleaner and cleaner.user else None,
                assignment.assignment_status if assignment else None,
                assignment.assigned_at if assignment else None,
                assignment.accepted_at if assignment else None,
                assignment.started_at if assignment else None,
                assignment.completed_at if assignment else None,
                payment.status if payment else "pending_collection",
                payment.payment_status if payment else "pending",
                payment.payment_type if payment else None,
                payment.collected_amount if payment else None,
                payment.cleaner_share if payment else None,
                payment.admin_share if payment else None,
                payment.cleaner_handover_status if payment else "pending",
                booking.special_instructions,
                booking.created_at,
                booking.updated_at,
            ]
        )
    return headers, rows


def _build_ratings_sheet(db: Session):
    ratings = (
        db.query(Rating)
        .options(
            selectinload(Rating.booking).selectinload(Booking.customer),
            selectinload(Rating.booking).selectinload(Booking.service_category),
            selectinload(Rating.booking)
            .selectinload(Booking.assignment)
            .selectinload(BookingAssignment.cleaner)
            .selectinload(CleanerProfile.user),
            selectinload(Rating.reviewer),
            selectinload(Rating.reviewee),
        )
        .order_by(Rating.created_at.desc())
        .all()
    )
    headers = [
        "Rating ID",
        "Booking ID",
        "Booking Reference",
        "Service",
        "Reviewer ID",
        "Reviewer Name",
        "Reviewer Role",
        "Reviewee ID",
        "Reviewee Name",
        "Rating",
        "Comment",
        "Customer Name",
        "Customer Phone",
        "Assigned Cleaner",
        "Created At",
    ]
    rows = []
    for rating in ratings:
        booking = rating.booking
        cleaner = booking.assignment.cleaner if booking and booking.assignment else None
        rows.append(
            [
                rating.id,
                rating.booking_id,
                booking.booking_reference if booking else None,
                booking.service_category.service_name if booking and booking.service_category else None,
                rating.reviewer_id,
                rating.reviewer.full_name if rating.reviewer else None,
                rating.reviewer_role,
                rating.reviewee_id,
                rating.reviewee.full_name if rating.reviewee else None,
                rating.rating,
                rating.comment,
                booking.customer.full_name if booking and booking.customer else None,
                booking.customer.phone if booking and booking.customer else None,
                cleaner.user.full_name if cleaner and cleaner.user else None,
                rating.created_at,
            ]
        )
    return headers, rows


def _build_payments_sheet(db: Session):
    payments = (
        db.query(Payment)
        .options(
            selectinload(Payment.booking).selectinload(Booking.customer),
            selectinload(Payment.booking).selectinload(Booking.service_category),
            selectinload(Payment.cleaner).selectinload(CleanerProfile.user),
            selectinload(Payment.split_admin),
        )
        .order_by(Payment.created_at.desc())
        .all()
    )
    headers = [
        "Payment ID",
        "Booking ID",
        "Booking Reference",
        "Customer ID",
        "Customer Name",
        "Customer Phone",
        "Service",
        "Payment Method",
        "Payment Type",
        "Transaction Reference",
        "Legacy Amount",
        "Collected Amount",
        "Payment Status",
        "Collection Status",
        "Collected By Cleaner",
        "Cleaner Profile ID",
        "Cleaner Name",
        "Collected At",
        "Cleaner Share",
        "Admin Share",
        "Split Updated By",
        "Split Updated At",
        "Payout Released",
        "Cleaner Handover Status",
        "Paid At",
        "Created At",
        "Updated At",
    ]
    rows = []
    for payment in payments:
        booking = payment.booking
        rows.append(
            [
                payment.id,
                payment.booking_id,
                booking.booking_reference if booking else None,
                payment.customer_id,
                booking.customer.full_name if booking and booking.customer else None,
                booking.customer.phone if booking and booking.customer else None,
                booking.service_category.service_name if booking and booking.service_category else None,
                payment.payment_method,
                payment.payment_type,
                payment.transaction_reference,
                payment.amount,
                payment.collected_amount,
                payment.payment_status,
                payment.status,
                payment.collected_by_cleaner,
                payment.collected_by,
                payment.cleaner.user.full_name if payment.cleaner and payment.cleaner.user else None,
                payment.collected_at,
                payment.cleaner_share,
                payment.admin_share,
                payment.split_admin.full_name if payment.split_admin else None,
                payment.split_updated_at,
                payment.payout_released,
                payment.cleaner_handover_status,
                payment.paid_at,
                payment.created_at,
                payment.updated_at,
            ]
        )
    return headers, rows


def _append_sheet(workbook: Workbook, title: str, headers: list[str], rows: list[list]):
    sheet = workbook.create_sheet(title=title[:31])
    sheet.append(headers)

    for row in rows:
        sheet.append([_excel_value(value) for value in row])

    header_fill = PatternFill("solid", fgColor="D9EAD3")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    _size_columns(sheet)


def _size_columns(sheet):
    for column_cells in sheet.columns:
        column = get_column_letter(column_cells[0].column)
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        sheet.column_dimensions[column].width = min(max(max_length + 2, 12), 42)


def _excel_value(value):
    if isinstance(value, UUID):
        return str(value)
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime):
        return value.replace(tzinfo=None).isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat(timespec="seconds")
    return value


def _role_names(user: User) -> str:
    roles: list[str] = []
    for user_role in user.user_roles:
        role = user_role.role
        if isinstance(role, Role) and role.role_name:
            roles.append(role.role_name)
    return ", ".join(roles)
